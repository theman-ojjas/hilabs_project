import ollama
import json
import pandas as pd
import email
from email.parser import BytesParser
from email import policy
from datetime import datetime
import openpyxl

def parse_email(eml_file):
    with open(eml_file, "rb") as f:
        msg = BytesParser(policy=policy.default).parse(f)
    
    # Extract body
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/plain":
                return part.get_content()
    return msg.get_content()

def create_excel(data_list, output_file):
    columns = [
        "Transaction Type (Add/Update/Term)",
        "Transaction Attribute",
        "Effective Date",
        "Term Date",
        "Term Reason",
        "Provider Name",
        "Provider NPI",
        "Provider Specialty",
        "State License",
        "Organization Name",
        "TIN",
        "Group NPI",
        "Complete Address",
        "Phone Number",
        "Fax Number",
        "PPG ID",
        "Line Of Business (Medicare/Commercial/Medical)"
    ]
    
    rows = []
    for data in data_list:
        rows.append({
            "Transaction Type (Add/Update/Term)": data["Transaction_Type"],
            "Transaction Attribute": data["Transaction_Attribute"],
            "Effective Date": data["Effective_Date"],
            "Term Date": data["Term_Date"],
            "Term Reason": data["Term_Reason"],
            "Provider Name": data["Provider_Name"],
            "Provider NPI": data["Provider_NPI"],
            "Provider Specialty": data["Provider_Specialty"],
            "State License": data["State_License"],
            "Organization Name": data["Organization_Name"],
            "TIN": data["TIN"],
            "Group NPI": data["Group_NPI"],
            "Complete Address": data["Complete_Address"],
            "Phone Number": data["Phone_Number"],
            "Fax Number": data["Fax_Number"],
            "PPG ID": data["PPG_ID"],
            "Line Of Business (Medicare/Commercial/Medical)": data["Line_Of_Business"]
        })
    
    df = pd.DataFrame(rows)
    df = df[columns]  # Ensure column order
    df.to_excel(output_file, index=False)

def process_provider_data(eml_file):
    client = ollama.Client()
    
    # Parse email content from .eml file
    email_body = parse_email(eml_file)
    
    # Stream response from LLM
    print("Generating response...")
    full_response = ""
    for chunk in client.generate(
        model='mario',
        prompt=email_body,
        stream=True
    ):
        chunk_text = chunk.get('response', '')
        print(chunk_text, end='', flush=True)
        full_response += chunk_text
    
    print("\n\nComplete response received.")
    
    # Parse the LLM response into JSON
    try:
        json_start = full_response.find('[')
        json_end = full_response.rfind(']') + 1
        if json_start >= 0 and json_end > json_start:
            json_str = full_response[json_start:json_end]
            data = json.loads(json_str)
            if not isinstance(data, list):
                data = [data]  # Convert single object to list
        else:
            raise ValueError("No JSON content found in response")
    except (json.JSONDecodeError, ValueError) as e:
        print(f"Error parsing response as JSON: {str(e)}")
        return None

    output_file = "Output.xlsx"
    
    # Check if file exists, if not create it
    import os
    if not os.path.exists(output_file):
        create_excel(data, output_file)
    else:
        append_to_excel(data, output_file)
    
    return output_file

def append_to_excel(data_list, output_file):
    max_retries = 3
    retries = 0
    
    while retries < max_retries:
        try:
            # Load existing workbook
            wb = openpyxl.load_workbook(output_file)
            ws = wb.active
            
            # Find the last row with data
            last_row = ws.max_row
            
            for data in data_list:
                row = last_row + 1
                ws.cell(row=row, column=1, value=str(data.get('Transaction_Type', '')))
                ws.cell(row=row, column=2, value=str(data.get('Transaction_Attribute', '')))
                ws.cell(row=row, column=3, value=str(data.get('Effective_Date', '')))
                ws.cell(row=row, column=4, value=str(data.get('Term_Date', '')))
                ws.cell(row=row, column=5, value=str(data.get('Term_Reason', '')))
                ws.cell(row=row, column=6, value=str(data.get('Provider_Name', '')))
                ws.cell(row=row, column=7, value=str(data.get('Provider_NPI', '')))
                ws.cell(row=row, column=8, value=str(data.get('Provider_Specialty', '')))
                ws.cell(row=row, column=9, value=str(data.get('State_License', '')))
                ws.cell(row=row, column=10, value=str(data.get('Organization_Name', '')))
                ws.cell(row=row, column=11, value=str(data.get('TIN', '')))
                ws.cell(row=row, column=12, value=str(data.get('Group_NPI', '')))
                ws.cell(row=row, column=13, value=str(data.get('Complete_Address', '')))
                ws.cell(row=row, column=14, value=str(data.get('Phone_Number', '')))
                ws.cell(row=row, column=15, value=str(data.get('Fax_Number', '')))
                ws.cell(row=row, column=16, value=str(data.get('PPG_ID', '')))
                ws.cell(row=row, column=17, value=str(data.get('Line_Of_Business', '')))
                last_row += 1
            
            # Save directly without temporary file
            wb.save(output_file)
            wb.close()
            return
            
        except PermissionError as e:
            print(f"Permission error (attempt {retries + 1}/{max_retries}): {str(e)}")
            retries += 1
            if retries < max_retries:
                import time
                time.sleep(1)
                continue
            raise
        except Exception as e:
            print(f"Error appending to Excel: {str(e)}")
            raise

# Example usage
if __name__ == "__main__":
    eml_file = "Sample-4.eml"  # Path to your .eml file
    output_file = process_provider_data(eml_file)
    print(f"Excel file created: {output_file}")